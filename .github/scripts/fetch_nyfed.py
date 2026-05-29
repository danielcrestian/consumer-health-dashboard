"""
Downloads the latest NY Fed Household Debt & Credit quarterly Excel file
and extracts "Percent of Balance 90+ Days Delinquent" by loan type.
Saves to data/delinquency.json.
"""
import io, json, os, re
from datetime import datetime, timezone

import openpyxl
import requests

DATA_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'delinquency.json')
BASE_URL  = 'https://www.newyorkfed.org/medialibrary/interactives/householdcredit/data/xls'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*',
}


# ── download ──────────────────────────────────────────────────────────────────

def quarter_sequence(start_year, start_q, steps=8):
    y, q = start_year, start_q
    for _ in range(steps):
        yield y, q
        q -= 1
        if q < 1:
            q, y = 4, y - 1


def download_latest():
    now = datetime.now(timezone.utc)
    for y, q in quarter_sequence(now.year, (now.month - 1) // 3 + 1):
        url = f'{BASE_URL}/hhd_c_report_{y}q{q}.xlsx'
        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            if r.status_code != 200:
                print(f'  HTTP {r.status_code} for {url}')
                continue
            if r.content[:4] != b'PK\x03\x04':
                print(f'  Skipping {url} — not a valid ZIP/XLSX')
                continue
            print(f'Downloaded: {url} ({len(r.content):,} bytes)')
            return r.content, url
        except requests.RequestException as e:
            print(f'  Request error for {url}: {e}')
    raise RuntimeError('Could not find NY Fed Excel file for last 8 quarters')


# ── sheet & data discovery ────────────────────────────────────────────────────

def is_data_sheet(ws):
    """Return True only for regular worksheets (not Chart sheets)."""
    return hasattr(ws, 'iter_rows')


def log_workbook_structure(wb):
    """Print full structure so we can diagnose parsing issues."""
    print(f'\n=== Workbook has {len(wb.sheetnames)} sheets ===')
    for name in wb.sheetnames:
        print(f'  Sheet: "{name}"')

    for name in wb.sheetnames:
        ws = wb[name]
        if not is_data_sheet(ws):
            continue          # skip Chartsheet objects
        print(f'\n--- First 8 rows of sheet "{name}" ---')
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 8:
                break
            cells = [str(c)[:25] if c is not None else '–' for c in row[:10]]
            print(f'  Row {i:2d}: {cells}')


def find_delinquency_sheet(wb):
    """
    Target 'Page 11 Data' — the NY Fed sheet for 90+ day delinquency.
    Falls back to scoring if the sheet name changes in future releases.
    """
    # Primary: known sheet name from NY Fed HHDC Excel format
    PREFERRED = ['Page 11 Data', 'page 11 data', 'Page11Data']
    for name in PREFERRED:
        if name in wb.sheetnames:
            print(f'Found target sheet: "{name}"')
            return wb[name]

    # Fallback: score data sheets by loan-type keyword matches
    print('WARNING: "Page 11 Data" not found — falling back to scoring')
    LOAN_KEYS = ['mortgage', 'auto', 'credit', 'student']
    DATE_RE   = re.compile(r'(19|20)\d{2}')
    best_name, best_score = None, -1

    for name in wb.sheetnames:
        ws = wb[name]
        if not is_data_sheet(ws):
            continue
        rows  = list(ws.iter_rows(min_row=1, max_row=60, values_only=True))
        score = 0
        for row in rows[:10]:
            text = ' '.join(str(c).lower() for c in row if c is not None)
            for kw in LOAN_KEYS:
                if kw in text:
                    score += 10
        for row in rows:
            first = str(row[0]).strip() if row[0] is not None else ''
            if DATE_RE.search(first):
                score += 1
        print(f'  Sheet "{name}" score: {score}')
        if score > best_score:
            best_score, best_name = score, name

    print(f'Selected: "{best_name}" (score {best_score})')
    return wb[best_name]


# ── date parsing ──────────────────────────────────────────────────────────────

def parse_date(raw):
    """Convert any NY Fed date format to YYYY-MM-DD (first day of quarter)."""
    if isinstance(raw, datetime):
        # Excel serial date decoded as datetime — snap to nearest quarter
        m = raw.month
        q = (m - 1) // 3 + 1
        return f'{raw.year}-{q * 3 - 2:02d}-01'
    if not isinstance(raw, str):
        return None
    raw = raw.strip()

    # "2025:Q1" or "2025:q1"
    m = re.match(r'(\d{4})[:\s\-]?[Qq](\d)', raw)
    if m:
        y, q = int(m.group(1)), int(m.group(2))
        return f'{y}-{q * 3 - 2:02d}-01'

    # "Q1 2025" or "Q1:2025"
    m = re.match(r'[Qq](\d)[:\s\-]+(\d{4})', raw)
    if m:
        q, y = int(m.group(1)), int(m.group(2))
        return f'{y}-{q * 3 - 2:02d}-01'

    # Bare year like "2003" — assume Q1
    m = re.match(r'^(\d{4})$', raw)
    if m:
        return f'{m.group(1)}-01-01'

    return None


# ── column detection ──────────────────────────────────────────────────────────

COLUMN_PATTERNS = {
    'mortgage':     ['mortgage'],
    'auto':         ['auto'],
    'credit_card':  ['credit card', 'credit_card', 'creditcard', 'cc'],
    'student_loan': ['student'],
}


def detect_columns(header_row):
    col_map = {}
    for ci, cell in enumerate(header_row):
        if cell is None:
            continue
        text = str(cell).lower().strip()
        for key, patterns in COLUMN_PATTERNS.items():
            if key not in col_map and any(p in text for p in patterns):
                col_map[key] = ci
                print(f'  Col {ci} → {key} ("{cell}")')
    return col_map


# ── main extraction ───────────────────────────────────────────────────────────

def extract_data(ws):
    data    = {k: [] for k in COLUMN_PATTERNS}
    col_map = {}
    header_found = False

    all_rows = list(ws.iter_rows(values_only=True))
    print(f'Total rows in sheet: {len(all_rows)}')

    for row_i, row in enumerate(all_rows):
        if not any(c is not None for c in row):
            continue

        # ── look for header row (first 20 rows only) ──
        if not header_found and row_i < 20:
            text = ' '.join(str(c).lower() for c in row if c is not None)
            if any(kw in text for kw in ['mortgage', 'auto', 'credit', 'student']):
                print(f'\nHeader row found at row {row_i}: {[str(c)[:20] for c in row[:10]]}')
                col_map = detect_columns(row)
                header_found = True
                continue

        if not header_found:
            continue

        # ── parse data row ──
        date_str = parse_date(row[0])
        if not date_str:
            continue

        for key, ci in col_map.items():
            if ci >= len(row) or row[ci] is None:
                continue
            try:
                val = float(row[ci])
                data[key].append({'date': date_str, 'value': round(val, 4)})
            except (ValueError, TypeError):
                pass

    for key, records in data.items():
        records.sort(key=lambda r: r['date'])
        print(f'  {key}: {len(records)} records', end='')
        if records:
            print(f'  [{records[0]["date"]} … {records[-1]["date"]}]')
        else:
            print()

    return data


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    content, source_url = download_latest()

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    log_workbook_structure(wb)

    print('\nScoring sheets to find delinquency data…')
    ws   = find_delinquency_sheet(wb)
    data = extract_data(ws)

    total = sum(len(v) for v in data.values())
    if total == 0:
        raise RuntimeError(
            'Extracted 0 records — sheet/column detection failed. '
            'Check the logs above to see sheet names and row contents.'
        )

    out = {
        'updated': datetime.now(timezone.utc).isoformat(),
        'source':  source_url,
        **data,
    }

    os.makedirs(os.path.dirname(DATA_PATH), exist_ok=True)
    with open(DATA_PATH, 'w') as f:
        json.dump(out, f, separators=(',', ':'))

    print(f'\nSaved {total} total records to {DATA_PATH}')


if __name__ == '__main__':
    main()
